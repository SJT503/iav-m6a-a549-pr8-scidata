# -*- coding: utf-8 -*-
"""03_enzymes.py — 抽取 m6A 调控酶完整清单（writer/reader/eraser）"""
import openpyxl
path = r"D:\IAV_m6A\Report\Report\A549-PR8的m6A数据的再分析 - 2022-4-13\m6A相关的酶\m6A酶表达情况xlsx.xlsx"
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
out = []
for ws in wb.worksheets:
    out.append(f"=== SHEET {ws.title} ({ws.max_row}x{ws.max_column}) ===")
    for ri, row in enumerate(ws.iter_rows(values_only=True)):
        gene = row[4] if len(row) > 4 else None
        fc_ge = row[1] if len(row) > 1 else None
        fc_m6a = row[2] if len(row) > 2 else None
        if gene and str(gene).strip() not in ("GeneSymbol", "", "None"):
            out.append(f"  {gene}\tFC(GE)={fc_ge}\tFC(m6A)={fc_m6a}")
wb.close()
OUT = r"D:\IAV_m6A\SciData_descriptor\results\m6a_enzymes.txt"
with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(out))
